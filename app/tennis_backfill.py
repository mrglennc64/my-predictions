"""Tennis backfill: tennis-data.co.uk results (2023-present, ATP + WTA)
-> Glicko-2 ratings, holdout-validated against Pinnacle closing odds,
written to tennis_ratings.

tennis-data names are 'Surname F.' — we key players as 'surname|f'. Live
Polymarket names ('First Surname') convert to the same key in tennis.py.

Usage: python -m app.tennis_backfill
"""
import io
import time
import unicodedata
from datetime import datetime, timezone

import requests
from openpyxl import load_workbook
from sqlalchemy import insert

from app import db
from src import devig
from src.models import glicko2

YEARS = (2023, 2024, 2025, 2026)
URLS = ([("atp", f"http://www.tennis-data.co.uk/{y}/{y}.xlsx") for y in YEARS]
        + [("wta", f"http://www.tennis-data.co.uk/{y}w/{y}.xlsx") for y in YEARS])
HOLDOUT_FROM = datetime(2026, 5, 1)


def key_from_result_name(name: str) -> str:
    """'Carreno Busta P.' -> 'carreno busta|p'"""
    nk = unicodedata.normalize("NFKD", name or "")
    nk = "".join(c for c in nk if not unicodedata.combining(c))
    parts = nk.replace(".", "").lower().split()
    if not parts:
        return ""
    if len(parts) >= 2 and len(parts[-1]) <= 2:
        return " ".join(parts[:-1]) + "|" + parts[-1][0]
    return " ".join(parts) + "|"


def load_matches():
    rows = []
    for tour, url in URLS:
        try:
            r = requests.get(url, timeout=120,
                             headers={"User-Agent": "Mozilla/5.0"})
            if not r.ok or not r.content[:2] == b"PK":
                print(f"  skip {url.rsplit('/', 1)[-1]} ({r.status_code})")
                continue
            ws = load_workbook(io.BytesIO(r.content), read_only=True).active
            header = None
            n = 0
            for row in ws.iter_rows(values_only=True):
                if header is None:
                    header = {str(v): i for i, v in enumerate(row)}
                    continue
                try:
                    date = row[header["Date"]]
                    winner = row[header["Winner"]]
                    loser = row[header["Loser"]]
                except (KeyError, IndexError):
                    continue
                if not (date and winner and loser):
                    continue
                psw = row[header["PSW"]] if "PSW" in header else None
                psl = row[header["PSL"]] if "PSL" in header else None
                rows.append((date, tour, str(winner), str(loser),
                             float(psw) if psw else None,
                             float(psl) if psl else None))
                n += 1
            print(f"  {tour} {url.rsplit('/', 1)[-1]}: {n} matches")
        except Exception as e:
            print(f"  error {url.rsplit('/', 1)[-1]}: {type(e).__name__} {e}")
    rows.sort(key=lambda x: x[0])
    return rows


def main():
    print("Downloading tennis-data.co.uk archives...")
    matches = load_matches()
    print(f"total: {len(matches)} matches")

    ratings: dict[str, glicko2.Rating] = {}
    meta: dict[str, tuple] = {}
    counts: dict[str, int] = {}
    m_sq = mk_sq = 0.0
    n_eval = n_mkt = m_correct = 0

    for date, tour, w_name, l_name, psw, psl in matches:
        ts = int(date.timestamp()) if hasattr(date, "timestamp") else int(time.time())
        wk, lk = key_from_result_name(w_name), key_from_result_name(l_name)
        if not wk or not lk:
            continue
        w = ratings.setdefault(wk, glicko2.Rating())
        l = ratings.setdefault(lk, glicko2.Rating())
        if (hasattr(date, "year") and date >= HOLDOUT_FROM
                and counts.get(wk, 0) >= 5 and counts.get(lk, 0) >= 5):
            p = glicko2.expected(glicko2.age(w, ts), glicko2.age(l, ts))
            m_sq += (p - 1.0) ** 2
            m_correct += p > 0.5
            n_eval += 1
            if psw and psl and psw > 1 and psl > 1:
                pw, _ = devig.devig_power([psw, psl])
                mk_sq += (pw - 1.0) ** 2
                n_mkt += 1
        glicko2.update(w, l, ts)
        meta[wk] = (w_name, tour)
        meta[lk] = (l_name, tour)
        counts[wk] = counts.get(wk, 0) + 1
        counts[lk] = counts.get(lk, 0) + 1

    print(f"\nHoldout since {HOLDOUT_FROM:%Y-%m-%d} (both players 5+ matches):")
    print(f"  model:    n={n_eval}  Brier={m_sq / max(n_eval, 1):.5f}  "
          f"accuracy={m_correct / max(n_eval, 1):.1%}")
    if n_mkt:
        print(f"  pinnacle: n={n_mkt}  Brier={mk_sq / n_mkt:.5f}  (de-vigged)")

    engine = db.init_db()
    now = int(datetime.now(timezone.utc).timestamp())
    with engine.begin() as conn:
        conn.execute(db.tennis_ratings.delete())
        payload = [{
            "player": k, "display": meta[k][0], "tour": meta[k][1],
            "rating": round(glicko2.age(r, now).rating, 1),
            "rd": round(r.rd, 1), "vol": round(r.vol, 5),
            "last_ts": r.last_ts, "matches": counts[k],
        } for k, r in ratings.items()]
        conn.execute(insert(db.tennis_ratings), payload)
    print(f"wrote {len(payload)} player ratings")
    for rating, name in sorted(((r.rating, meta[k][0])
                                for k, r in ratings.items()
                                if counts[k] >= 20), reverse=True)[:8]:
        print(f"  {name:28} {rating:7.1f}")


if __name__ == "__main__":
    main()
